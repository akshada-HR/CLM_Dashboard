# -*- coding: utf-8 -*-
"""
=============================================================================
CONTRACT LABOUR COMPLIANCE CONTROL TOWER
=============================================================================
Single-file Streamlit application.
Developed for: HR Contractor Cell
Purpose      : Upload CLM export -> monitor vendor compliance, track issues,
               generate management reports. Supports multi-month selection.
=============================================================================
"""

import io
import re
import sqlite3
import datetime

import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# CONFIGURATION
# -----------------------------------------------------------------------------

APP_TITLE = "Contract Labour Compliance Control Tower"
DB_PATH   = "compliance_control_tower.db"

MONTH_ORDER = {
    "Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
    "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12
}

REQUIRED_COLUMNS = [
    "Region","Location","Vendor Name","V Code",
    "No of Employees","Active gatepass","Difference",
    "Wage Compliance","PF Compliance","ESIC Compliance",
    "Reason for Non-Compliance"
]

STATUS_OPTIONS = [
    "New",
    "Waiting for Vendor Documents",
    "Vendor Responded",
    "Under HR Review",
    "Under Contractor Cell Review",
    "PO Pending",
    "Medical Pending",
    "Safety Training Pending",
    "Gate Pass Under Process",
    "Escalated",
    "Resolved",
    "Closed",
]

OWNER_OPTIONS = [
    "HR","Contractor Cell","Vendor",
    "Procurement","Security","Medical",
    "Safety","Finance",
]

HEALTH_BANDS = [
    (100, 100, "Healthy",          "#27ae60"),
    (75,   99, "Minor Issues",     "#f39c12"),
    (50,   74, "Needs Attention",  "#e67e22"),
    (0,    49, "Critical",         "#e74c3c"),
]

# -----------------------------------------------------------------------------
# PAGE CONFIG
# -----------------------------------------------------------------------------

st.set_page_config(
    page_title=APP_TITLE,
    page_icon="CT",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -----------------------------------------------------------------------------
# CSS  -- all text forced to readable colours on their backgrounds
# -----------------------------------------------------------------------------

st.markdown("""
<style>
    /* ── global ── */
    .stApp { background-color: #f0f2f6; color: #1a1a1a; }

    /* ── all default Streamlit text elements ── */
    p, span, label, div, li, td, th,
    .stMarkdown, .stText,
    [data-testid="stMetricValue"],
    [data-testid="stMetricLabel"],
    [data-testid="stCaptionContainer"],
    [data-testid="stExpander"] summary,
    [data-testid="stDataFrame"] { color: #1a1a1a !important; }

    h1, h2, h3, h4 { color: #1a1a1a !important; font-weight: 700; }

    /* ── sidebar  -- dark bg, white text ── */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a3c5e 0%, #2c6fad 100%);
    }
    [data-testid="stSidebar"] *,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] caption,
    [data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
        color: #ffffff !important;
    }

    /* ── KPI cards  -- white bg, dark text ── */
    .kpi-card {
        background: #ffffff;
        border-radius: 10px;
        padding: 20px 16px;
        text-align: center;
        border-left: 5px solid #2c6fad;
        box-shadow: 0 2px 8px rgba(0,0,0,.10);
    }
    .kpi-card.red    { border-left-color: #c0392b; }
    .kpi-card.green  { border-left-color: #1e8449; }
    .kpi-card.amber  { border-left-color: #d68910; }
    .kpi-card.purple { border-left-color: #6c3483; }
    .kpi-value { font-size: 2rem; font-weight: 800; color: #1a1a1a !important; }
    .kpi-label { font-size: .82rem; color: #333333 !important; margin-top: 6px; font-weight: 600; }

    /* ── section header  -- dark bg, white text ── */
    .section-header {
        background: linear-gradient(90deg, #1a3c5e, #2c6fad);
        color: #ffffff !important;
        padding: 10px 18px;
        border-radius: 6px;
        font-weight: 700;
        font-size: 1rem;
        margin-bottom: 14px;
    }

    /* ── month selector banner ── */
    .month-banner {
        background: #1a3c5e;
        color: #ffffff !important;
        padding: 8px 16px;
        border-radius: 6px;
        font-weight: 600;
        margin-bottom: 10px;
        font-size: 0.95rem;
    }

    /* ── info / success / warning boxes ── */
    [data-testid="stAlert"] p { color: #1a1a1a !important; }

    /* ── tabs ── */
    [data-baseweb="tab-list"] { background: #ffffff; border-radius: 8px; }
    [data-baseweb="tab"] { font-weight: 700; color: #1a1a1a !important; }
    [data-baseweb="tab"][aria-selected="true"] { color: #1a3c5e !important; }

    /* ── dataframe ── */
    [data-testid="stDataFrame"] td,
    [data-testid="stDataFrame"] th { color: #1a1a1a !important; }

    /* ── selectbox / text input labels ── */
    [data-testid="stSelectbox"] label,
    [data-testid="stTextInput"] label,
    [data-testid="stTextArea"] label { color: #1a1a1a !important; font-weight: 600; }

    /* ── filter dropdown boxes: dark navy bg, white text (Region/Status/Priority, search) ── */
    [data-testid="stSelectbox"] > div > div,
    [data-testid="stSelectbox"] [data-baseweb="select"] > div,
    [data-testid="stTextInput"] input {
        background-color: #1a3c5e !important;
        color: #ffffff !important;
    }
    [data-testid="stSelectbox"] svg { fill: #ffffff !important; }

    /* the open dropdown menu renders in a portal -- target it directly */
    ul[data-testid="stSelectboxVirtualDropdown"],
    div[data-baseweb="popover"] ul,
    div[data-baseweb="menu"],
    div[role="listbox"] {
        background-color: #1a3c5e !important;
    }
    ul[data-testid="stSelectboxVirtualDropdown"] li,
    div[data-baseweb="popover"] li,
    div[data-baseweb="menu"] li,
    div[role="listbox"] li,
    div[role="listbox"] [role="option"],
    ul[data-testid="stSelectboxVirtualDropdown"] li *,
    div[data-baseweb="popover"] li *,
    div[data-baseweb="menu"] li *,
    div[role="listbox"] li *,
    div[role="listbox"] [role="option"] * {
        background-color: #1a3c5e !important;
        color: #ffffff !important;
    }
    ul[data-testid="stSelectboxVirtualDropdown"] li:hover,
    div[data-baseweb="popover"] li:hover,
    div[data-baseweb="menu"] li:hover,
    div[role="listbox"] li:hover,
    div[role="listbox"] [role="option"]:hover {
        background-color: #2c6fad !important;
    }

    /* ── expander header ── */
    [data-testid="stExpander"] summary p { color: #1a1a1a !important; font-weight: 600; }

    /* ── upload zone ── */
    [data-testid="stFileUploader"] { background: #ffffff; border-radius: 10px; padding: 10px; }

    /* ── buttons (download / regular) -- dark navy bg, white text ── */
    [data-testid="stDownloadButton"] button,
    [data-testid="stBaseButton-secondary"],
    [data-testid="stBaseButton-primary"],
    .stButton button,
    .stDownloadButton button {
        background-color: #1a3c5e !important;
        color: #ffffff !important;
        border: 1px solid #1a3c5e !important;
        font-weight: 600;
    }
    [data-testid="stDownloadButton"] button *,
    .stButton button *,
    .stDownloadButton button * {
        color: #ffffff !important;
    }
    [data-testid="stDownloadButton"] button:hover,
    [data-testid="stBaseButton-secondary"]:hover,
    [data-testid="stBaseButton-primary"]:hover,
    .stButton button:hover,
    .stDownloadButton button:hover {
        background-color: #2c6fad !important;
        color: #ffffff !important;
        border: 1px solid #2c6fad !important;
    }

    /* ── st.code() blocks -- dark navy bg, white text ── */
    [data-testid="stCodeBlock"],
    [data-testid="stCodeBlock"] pre,
    [data-testid="stCodeBlock"] code,
    [data-testid="stCodeBlock"] pre code,
    .stCodeBlock,
    .stCodeBlock pre,
    .stCodeBlock code {
        background-color: #1a3c5e !important;
        color-scheme: dark;
    }
    [data-testid="stCodeBlock"] pre code span,
    [data-testid="stCodeBlock"] code span,
    [data-testid="stCodeBlock"] span,
    [data-testid="stCodeBlock"] pre,
    [data-testid="stCodeBlock"] code,
    [data-testid="stCodeBlock"] *,
    .stCodeBlock span,
    .stCodeBlock * {
        color: #ffffff !important;
        background-color: transparent !important;
        -webkit-text-fill-color: #ffffff !important;
    }
</style>
""", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# DATABASE LAYER
# -----------------------------------------------------------------------------

def get_db():
    """Return SQLite connection; create tables if not present."""
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    cur  = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS TrackingData (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            report_month   TEXT,
            vendor_name    TEXT,
            current_status TEXT DEFAULT 'New',
            owner          TEXT DEFAULT 'HR',
            remarks        TEXT DEFAULT '',
            last_updated   TEXT,
            UNIQUE(report_month, vendor_name)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS AuditLog (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            ts           TEXT,
            report_month TEXT,
            vendor_name  TEXT,
            field        TEXT,
            old_value    TEXT,
            new_value    TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS UploadHistory (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            report_month TEXT UNIQUE,
            uploaded_at  TEXT,
            row_count    INTEGER,
            vendor_count INTEGER
        )
    """)

    # Store raw vendor data per month for multi-month replay
    cur.execute("""
        CREATE TABLE IF NOT EXISTS MonthlyVendorData (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            report_month TEXT,
            vendor_json  TEXT,
            UNIQUE(report_month)
        )
    """)

    conn.commit()
    return conn


def upsert_tracking(conn, report_month, vendor, status, owner, remarks):
    """Insert or update tracking record and write audit log on change."""
    today = datetime.date.today().isoformat()
    cur   = conn.cursor()

    cur.execute(
        "SELECT current_status, owner, remarks FROM TrackingData "
        "WHERE report_month=? AND vendor_name=?",
        (report_month, vendor)
    )
    existing = cur.fetchone()

    if existing:
        old_status, old_owner, old_remarks = existing
        for field, old, new in [
            ("current_status", old_status,  status),
            ("owner",          old_owner,   owner),
            ("remarks",        old_remarks, remarks),
        ]:
            if old != new:
                cur.execute(
                    "INSERT INTO AuditLog"
                    "(ts,report_month,vendor_name,field,old_value,new_value) "
                    "VALUES(?,?,?,?,?,?)",
                    (today, report_month, vendor, field, old, new)
                )
        cur.execute(
            "UPDATE TrackingData "
            "SET current_status=?, owner=?, remarks=?, last_updated=? "
            "WHERE report_month=? AND vendor_name=?",
            (status, owner, remarks, today, report_month, vendor)
        )
    else:
        cur.execute(
            "INSERT INTO TrackingData"
            "(report_month,vendor_name,current_status,owner,remarks,last_updated) "
            "VALUES(?,?,?,?,?,?)",
            (report_month, vendor, status, owner, remarks, today)
        )

    conn.commit()


def get_tracking(conn, report_month):
    return pd.read_sql_query(
        "SELECT vendor_name, current_status, owner, remarks, last_updated "
        "FROM TrackingData WHERE report_month=?",
        conn, params=(report_month,)
    )


def save_monthly_vendor_data(conn, report_month, vendor_df):
    """Persist vendor-level data for a month so it can be reloaded later."""
    import json
    vendor_json = vendor_df.to_json(orient="records")
    conn.execute(
        "INSERT OR REPLACE INTO MonthlyVendorData(report_month, vendor_json) VALUES(?,?)",
        (report_month, vendor_json)
    )
    conn.commit()


def load_monthly_vendor_data(conn, report_month):
    """Load persisted vendor data for a month. Returns None if not found."""
    import json
    cur = conn.cursor()
    cur.execute(
        "SELECT vendor_json FROM MonthlyVendorData WHERE report_month=?",
        (report_month,)
    )
    row = cur.fetchone()
    if row:
        return pd.read_json(io.StringIO(row[0]), orient="records")
    return None


def get_all_stored_months(conn):
    """Return sorted list of all months that have stored vendor data."""
    cur = conn.cursor()
    cur.execute("SELECT report_month FROM MonthlyVendorData ORDER BY report_month")
    rows = cur.fetchall()
    months = [r[0] for r in rows]
    # Sort by calendar order
    def month_sort_key(m):
        parts = m.strip().split()
        if len(parts) == 2:
            return (int(parts[1]), MONTH_ORDER.get(parts[0][:3], 0))
        return (9999, 0)
    return sorted(months, key=month_sort_key)


def log_upload(conn, report_month, row_count, vendor_count):
    conn.execute(
        "INSERT OR REPLACE INTO UploadHistory"
        "(report_month, uploaded_at, row_count, vendor_count) "
        "VALUES(?,?,?,?)",
        (report_month, datetime.datetime.now().isoformat(), row_count, vendor_count)
    )
    conn.commit()


def get_upload_history(conn):
    return pd.read_sql_query(
        "SELECT report_month, uploaded_at, row_count, vendor_count "
        "FROM UploadHistory ORDER BY uploaded_at DESC",
        conn
    )


# -----------------------------------------------------------------------------
# COLUMN FUZZY MATCHING
# -----------------------------------------------------------------------------

def _canonical(col_name):
    """
    Canonical key: lowercase, strip all spaces/hyphens/underscores.
    e.g. "Reason  for Non- Compliance" -> "reasonfornoncompliance"
    """
    s = str(col_name).lower()
    s = re.sub(r"[\s\-_/]+", "", s)
    return s


_REQUIRED_CANONICAL = {_canonical(c): c for c in REQUIRED_COLUMNS}


def normalise_columns(df):
    """Fuzzy-rename columns to match required names regardless of spacing/hyphens."""
    df = df.copy()
    df.columns = df.columns.str.strip()
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]

    rename_map = {}
    for actual_col in df.columns:
        canon = _canonical(actual_col)
        if canon in _REQUIRED_CANONICAL:
            target = _REQUIRED_CANONICAL[canon]
            if actual_col != target:
                rename_map[actual_col] = target

    if rename_map:
        df.rename(columns=rename_map, inplace=True)

    return df


def validate_columns(df):
    return [c for c in REQUIRED_COLUMNS if c not in df.columns]


# -----------------------------------------------------------------------------
# FILE PARSING  -- now returns ALL months' data, not just latest
# -----------------------------------------------------------------------------

def parse_upload(file):
    """
    Parse uploaded file.
    For Excel: reads EVERY monthly sheet and returns a dict {month: dataframe}.
    For CSV:   returns a single-entry dict {"CSV Upload": dataframe}.
    Also returns sorted list of month names.
    """
    name = file.name.lower()

    if name.endswith(".csv"):
        df = pd.read_csv(file)
        df = normalise_columns(df)
        return {"CSV Upload": df}, ["CSV Upload"]

    xls        = pd.ExcelFile(file, engine='openpyxl')
    all_months = [
        s for s in xls.sheet_names
        if re.match(r"^[A-Za-z]{3}\s+\d{2}$", s.strip())
    ]
    if not all_months:
        st.error(
            "No monthly sheets detected. "
            "Expected sheet names like Jan 26, Feb 26 ... Dec 26."
        )
        st.stop()

    sorted_months = sorted(
        all_months,
        key=lambda x: MONTH_ORDER.get(x.strip()[:3], 0)
    )

    month_data = {}
    for month in sorted_months:
        df = pd.read_excel(file, sheet_name=month, engine='openpyxl')
        df = normalise_columns(df)
        month_data[month] = df

    return month_data, sorted_months


# -----------------------------------------------------------------------------
# BUSINESS LOGIC
# -----------------------------------------------------------------------------

def get_difference(row):
    """Single source of truth for the gatepass/headcount gap.
    Avoids the two callers (health score vs issue list) disagreeing
    on what a missing/NaN Difference value should mean."""
    diff = row.get("Difference", 0)
    if diff is None or (isinstance(diff, float) and pd.isna(diff)):
        return 0
    return diff


def compute_health_score(row):
    score = 0
    if str(row.get("Wage Compliance","")).strip().lower() == "complied":
        score += 25
    if str(row.get("PF Compliance","")).strip().lower() == "complied":
        score += 25
    if str(row.get("ESIC Compliance","")).strip().lower() == "complied":
        score += 25
    if get_difference(row) == 0:
        score += 25
    return score


def health_label(score):
    for lo, hi, label, _ in HEALTH_BANDS:
        if lo <= score <= hi:
            return label
    return "Unknown"


def get_issues(row):
    issues = []
    if str(row.get("Wage Compliance","")).strip().lower() != "complied":
        issues.append("Wage Non-Compliant")
    if str(row.get("PF Compliance","")).strip().lower() != "complied":
        issues.append("PF Non-Compliant")
    if str(row.get("ESIC Compliance","")).strip().lower() != "complied":
        issues.append("ESIC Non-Compliant")
    diff = get_difference(row)
    if diff != 0:
        issues.append(f"Gatepass Gap ({int(diff)} employees)")
    return issues


def get_priority(issues):
    if not issues:
        return "-"
    if any(any(k in iss for k in ["Wage","PF","ESIC"]) for iss in issues):
        return "Critical"
    if "Gatepass" in " ".join(issues):
        return "High"
    return "Medium"


def get_recommendations(row):
    recs = []
    if str(row.get("PF Compliance","")).strip().lower() != "complied":
        recs.append("Verify PF challan and payment records immediately")
    if str(row.get("ESIC Compliance","")).strip().lower() != "complied":
        recs.append("Verify ESIC contributions and challan")
    if str(row.get("Wage Compliance","")).strip().lower() != "complied":
        recs.append("Contact vendor - wage payment evidence required")
    if get_difference(row) != 0:
        recs.append("Reconcile gatepass vs deployment headcount")
    return " | ".join(recs) if recs else "No Action Required"


def build_vendor_df(raw):
    """Aggregate raw rows to one row per vendor."""
    agg_cols = {
        "Region"                   : "first",
        "Location"                 : "first",
        "No of Employees"          : "sum",
        "Active gatepass"          : "sum",
        "Difference"               : "sum",
        "Wage Compliance"          : "first",
        "PF Compliance"            : "first",
        "ESIC Compliance"          : "first",
        "Reason for Non-Compliance": lambda x: " | ".join(
            [str(v) for v in x.dropna().unique() if str(v).strip()]
        ),
    }
    # If the CLM export has its own "City" column, aggregate it directly.
    # Otherwise City is derived later from Location as a fallback.
    if "City" in raw.columns:
        agg_cols["City"] = "first"

    agg_cols = {k: v for k, v in agg_cols.items() if k in raw.columns}
    vendor   = raw.groupby("Vendor Name", as_index=False).agg(agg_cols)

    if "V Code" in raw.columns:
        vc     = raw.groupby("Vendor Name")["V Code"].first().reset_index()
        vendor = vendor.merge(vc, on="Vendor Name", how="left")

    # Ensure a City column always exists. If the upload didn't include a
    # dedicated City field, fall back to Location so every table/report
    # still has a city value to show.
    if "City" not in vendor.columns:
        vendor["City"] = vendor["Location"] if "Location" in vendor.columns else ""

    # Defensive recompute: don't blindly trust the upstream "Difference"
    # column. It should represent the headcount gap between employees on
    # the books and employees with an active gatepass. Recomputing it
    # here protects against upstream formula drift (e.g. someone editing
    # the source sheet) and floors it at 0 -- a vendor having MORE active
    # gatepasses than employees isn't a compliance gap in the same sense
    # as missing gatepasses, so it shouldn't be flagged as one.
    if "No of Employees" in vendor.columns and "Active gatepass" in vendor.columns:
        vendor["Difference"] = (
            vendor["No of Employees"] - vendor["Active gatepass"]
        ).clip(lower=0)

    vendor["Health Score"]       = vendor.apply(compute_health_score, axis=1)
    vendor["Status"]             = vendor["Health Score"].apply(health_label)
    vendor["Issues"]             = vendor.apply(
        lambda r: " | ".join(get_issues(r)) or "None", axis=1
    )
    vendor["Priority"]           = vendor.apply(
        lambda r: get_priority(get_issues(r)), axis=1
    )
    vendor["Recommended Action"] = vendor.apply(get_recommendations, axis=1)
    return vendor


def compute_kpis(vendor):
    total    = len(vendor)
    comp     = int((vendor["Health Score"] == 100).sum())
    non      = total - comp
    emp      = int(vendor["No of Employees"].sum())
    risk     = int(vendor.loc[vendor["Health Score"] < 100, "No of Employees"].sum())
    pct      = round((comp / total) * 100, 1) if total else 0.0
    critical = int((vendor["Health Score"] < 50).sum())
    return dict(total=total, comp=comp, non=non, emp=emp,
                risk=risk, pct=pct, critical=critical)


def compute_aging(last_updated_str):
    val = str(last_updated_str).strip()
    if val in ("", "nan", "None"):
        return "Not updated"
    try:
        lu = datetime.date.fromisoformat(val)
        d  = (datetime.date.today() - lu).days
        if d == 0:  return "Today"
        if d == 1:  return "1 day"
        if d < 7:   return f"{d} days"
        if d < 30:  return f"{d // 7} week{'s' if d // 7 > 1 else ''}"
        return f"{d // 30} month{'s' if d // 30 > 1 else ''}"
    except Exception:
        return "-"


# -----------------------------------------------------------------------------
# CHARTS
# -----------------------------------------------------------------------------

def chart_donut(comp, non, key_suffix=""):
    fig = go.Figure(go.Pie(
        labels=["Compliant","Non-Compliant"],
        values=[comp, non],
        hole=.60,
        marker_colors=["#27ae60","#e74c3c"],
        textinfo="percent+label",
        hovertemplate="%{label}: %{value} vendors<extra></extra>",
    ))
    fig.update_layout(
        title="Vendor Compliance Split",
        height=320,
        margin=dict(t=40,b=10,l=10,r=10),
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="#1a1a1a", size=12),
        title_font=dict(color="#1a1a1a", size=14),
        legend=dict(font=dict(color="#1a1a1a")),
        annotations=[dict(text=f"{comp+non}<br>Vendors",
                          x=.5, y=.5, font_size=14, showarrow=False,
                          font_color="#1a1a1a")]
    )
    return fig


def chart_region_bar(vendor, key_suffix=""):
    reg = (vendor[vendor["Health Score"] < 100]
           .groupby("Region").size()
           .reset_index(name="Non-Compliant Vendors"))
    fig = px.bar(reg, x="Region", y="Non-Compliant Vendors",
                 color="Non-Compliant Vendors",
                 color_continuous_scale=["#f39c12","#e74c3c"],
                 title="Region-wise Non-Compliant Vendors",
                 text="Non-Compliant Vendors")
    fig.update_traces(textposition="outside", textfont_color="#1a1a1a")
    fig.update_layout(height=320, margin=dict(t=40,b=10,l=10,r=10),
                      coloraxis_showscale=False, paper_bgcolor="white",
                      plot_bgcolor="white",
                      font=dict(color="#1a1a1a", size=12),
                      title_font=dict(color="#1a1a1a", size=14),
                      xaxis=dict(tickfont=dict(color="#1a1a1a"), title_font=dict(color="#1a1a1a")),
                      yaxis=dict(tickfont=dict(color="#1a1a1a"), title_font=dict(color="#1a1a1a")))
    return fig


def chart_employees_at_risk(vendor):
    emp = (vendor[vendor["Health Score"] < 100]
           .groupby("Region")["No of Employees"]
           .sum().reset_index(name="Employees at Risk"))
    fig = px.bar(emp, x="Region", y="Employees at Risk",
                 color="Employees at Risk",
                 color_continuous_scale=["#f39c12","#e74c3c"],
                 title="Employees at Risk by Region",
                 text="Employees at Risk")
    fig.update_traces(textposition="outside", textfont_color="#1a1a1a")
    fig.update_layout(height=320, margin=dict(t=40,b=10,l=10,r=10),
                      coloraxis_showscale=False, paper_bgcolor="white",
                      plot_bgcolor="white",
                      font=dict(color="#1a1a1a", size=12),
                      title_font=dict(color="#1a1a1a", size=14),
                      xaxis=dict(tickfont=dict(color="#1a1a1a"), title_font=dict(color="#1a1a1a")),
                      yaxis=dict(tickfont=dict(color="#1a1a1a"), title_font=dict(color="#1a1a1a")))
    return fig


def chart_root_cause(vendor):
    reasons = []
    for r in vendor["Reason for Non-Compliance"].dropna():
        for part in str(r).split("|"):
            part = part.strip()
            if part and part.lower() not in ("nan",""):
                reasons.append(part)
    if not reasons:
        return None
    rc = pd.Series(reasons).value_counts().reset_index()
    rc.columns = ["Reason","Count"]
    fig = px.bar(rc.head(10), x="Count", y="Reason", orientation="h",
                 title="Root Cause Analysis - Top Non-Compliance Reasons",
                 color="Count",
                 color_continuous_scale=["#f39c12","#e74c3c"])
    fig.update_layout(height=350, margin=dict(t=40,b=10,l=10,r=10),
                      coloraxis_showscale=False,
                      yaxis=dict(autorange="reversed", tickfont=dict(color="#1a1a1a"), title_font=dict(color="#1a1a1a")),
                      xaxis=dict(tickfont=dict(color="#1a1a1a"), title_font=dict(color="#1a1a1a")),
                      paper_bgcolor="white",
                      plot_bgcolor="white",
                      font=dict(color="#1a1a1a", size=12),
                      title_font=dict(color="#1a1a1a", size=14))
    return fig


def chart_priority_breakdown(vendor):
    nc = vendor[vendor["Health Score"] < 100]
    if nc.empty:
        return None
    pc = nc["Priority"].value_counts().reset_index()
    pc.columns = ["Priority","Count"]
    colour_map = {"Critical":"#e74c3c","High":"#e67e22","Medium":"#f39c12"}
    fig = px.bar(pc, x="Priority", y="Count",
                 color="Priority", color_discrete_map=colour_map,
                 title="Non-Compliant Vendors by Priority",
                 text="Count")
    fig.update_traces(textposition="outside", textfont_color="#1a1a1a")
    fig.update_layout(height=300, showlegend=False,
                      margin=dict(t=40,b=10,l=10,r=10),
                      paper_bgcolor="white", plot_bgcolor="white",
                      font=dict(color="#1a1a1a", size=12),
                      title_font=dict(color="#1a1a1a", size=14),
                      xaxis=dict(tickfont=dict(color="#1a1a1a")),
                      yaxis=dict(tickfont=dict(color="#1a1a1a")))
    return fig


def chart_component_compliance(vendor):
    total  = len(vendor)
    wage_c = (vendor["Wage Compliance"].str.lower().str.strip()=="complied").sum()
    pf_c   = (vendor["PF Compliance"].str.lower().str.strip()=="complied").sum()
    esic_c = (vendor["ESIC Compliance"].str.lower().str.strip()=="complied").sum()
    diff_c = (vendor["Difference"]==0).sum()

    comp_data = pd.DataFrame({
        "Component": ["Wage","PF","ESIC","Gatepass"],
        "Compliant": [wage_c, pf_c, esic_c, diff_c],
        "Non-Compliant": [total-wage_c, total-pf_c, total-esic_c, total-diff_c],
    })
    fig = go.Figure(data=[
        go.Bar(name="Compliant",     x=comp_data["Component"],
               y=comp_data["Compliant"],     marker_color="#27ae60"),
        go.Bar(name="Non-Compliant", x=comp_data["Component"],
               y=comp_data["Non-Compliant"], marker_color="#e74c3c"),
    ])
    fig.update_layout(barmode="group", height=300,
                      title="Compliance by Component",
                      margin=dict(t=40,b=10,l=10,r=10),
                      paper_bgcolor="white", plot_bgcolor="white",
                      font=dict(color="#1a1a1a", size=12),
                      title_font=dict(color="#1a1a1a", size=14),
                      legend=dict(font=dict(color="#1a1a1a")),
                      xaxis=dict(tickfont=dict(color="#1a1a1a")),
                      yaxis=dict(tickfont=dict(color="#1a1a1a")))
    return fig


# -----------------------------------------------------------------------------
# EXCEL REPORT BUILDERS
# -----------------------------------------------------------------------------

def _hdr_style(ws, row, cols, fill_hex="1a3c5e", font_hex="FFFFFF"):
    fill  = PatternFill("solid", fgColor=fill_hex)
    font  = Font(bold=True, color=font_hex, size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    for c in range(1, cols+1):
        cell           = ws.cell(row=row, column=c)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
        cell.border    = thin


def _alt_rows(ws, start_row, end_row, cols, even_hex="EAF0F6"):
    fill_even = PatternFill("solid", fgColor=even_hex)
    for r in range(start_row, end_row+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            if r % 2 == 0:
                cell.fill = fill_even


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, min_w), max_w
        )


def build_executive_summary_sheet(wb, kpis, report_month):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Contract Labour Compliance - Executive Summary  |  {report_month}"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1a3c5e")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Generated: {datetime.date.today().strftime('%d %b %Y')}"
    ws["A2"].font       = Font(italic=True, size=9, color="000000")
    ws.row_dimensions[2].height = 18

    headers = ["Total Vendors","Total Employees","Compliant","Non-Compliant",
               "Employees at Risk","Compliance %","Critical Vendors"]
    values  = [kpis["total"], kpis["emp"], kpis["comp"], kpis["non"],
               kpis["risk"], f"{kpis['pct']}%", kpis["critical"]]
    colours = ["2c6fad","16a085","27ae60","e74c3c","e67e22","8e44ad","c0392b"]

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 36
    for i, (h, v, col) in enumerate(zip(headers, values, colours), start=1):
        c_h = ws.cell(row=4, column=i, value=h)
        c_v = ws.cell(row=5, column=i, value=v)
        c_h.fill      = PatternFill("solid", fgColor=col)
        c_h.font      = Font(bold=True, color="FFFFFF", size=9)
        c_h.alignment = Alignment(horizontal="center", vertical="center")
        c_v.font      = Font(bold=True, size=16, color=col)
        c_v.alignment = Alignment(horizontal="center", vertical="center")

    _auto_width(ws)
    return ws


def build_vendor_status_sheet(wb, vendor):
    ws = wb.create_sheet("Vendor Status")
    ws.sheet_view.showGridLines = False

    cols_wanted = ["Vendor Name","V Code","Region","Location","City",
                   "No of Employees","Health Score","Status",
                   "Issues","Priority","Recommended Action"]
    cols_present = [c for c in cols_wanted if c in vendor.columns]

    for ci, h in enumerate(cols_present, 1):
        ws.cell(row=1, column=ci, value=h)
    _hdr_style(ws, 1, len(cols_present))
    ws.row_dimensions[1].height = 22

    status_colours = {"Healthy":"27ae60","Minor Issues":"f39c12",
                      "Needs Attention":"e67e22","Critical":"e74c3c"}
    for ri, (_, row) in enumerate(vendor[cols_present].iterrows(), start=2):
        for ci, col in enumerate(cols_present, 1):
            cell = ws.cell(row=ri, column=ci, value=row[col])
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            if col == "Status":
                hex_c     = status_colours.get(str(row[col]), "999999")
                cell.fill = PatternFill("solid", fgColor=hex_c)
                cell.font = Font(bold=True, color="FFFFFF")
            if col == "Health Score":
                cell.font = Font(bold=True, size=11)
        if ri % 2 == 0:
            for ci in range(1, len(cols_present)+1):
                c2  = ws.cell(row=ri, column=ci)
                rgb = c2.fill.fgColor.rgb if c2.fill.fgColor else ""
                if rgb in ("00000000","FFFFFFFF",""):
                    c2.fill = PatternFill("solid", fgColor="EAF0F6")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_present))}1"
    _auto_width(ws)
    return ws


def build_action_center_sheet(wb, action_df):
    ws = wb.create_sheet("Action Center")
    ws.sheet_view.showGridLines = False

    cols_wanted = ["Vendor Name","Region","City","Issues","Priority",
                   "Current Status","Owner","Remarks","Last Updated",
                   "Status Aging","Recommended Action"]
    cols_present = [c for c in cols_wanted if c in action_df.columns]

    for ci, h in enumerate(cols_present, 1):
        ws.cell(row=1, column=ci, value=h)
    _hdr_style(ws, 1, len(cols_present), fill_hex="c0392b")
    ws.row_dimensions[1].height = 22

    priority_font = {
        "Critical": Font(bold=True, color="c0392b"),
        "High"    : Font(bold=True, color="e67e22"),
        "Medium"  : Font(bold=True, color="d68910"),
    }
    for ri, (_, row) in enumerate(action_df[cols_present].iterrows(), start=2):
        for ci, col in enumerate(cols_present, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col,""))
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            if col == "Priority":
                cell.font = priority_font.get(str(row.get(col,"")), Font(size=10))
        if ri % 2 == 0:
            for ci in range(1, len(cols_present)+1):
                ws.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="FDEDEC")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_present))}1"
    _auto_width(ws)
    return ws


def build_complete_report(vendor, kpis, action_df, report_month):
    wb = Workbook()
    wb.remove(wb.active)
    build_executive_summary_sheet(wb, kpis, report_month)
    build_vendor_status_sheet(wb, vendor)
    build_action_center_sheet(wb, action_df)

    ws_raw = wb.create_sheet("Raw Data")
    ws_raw.sheet_view.showGridLines = False
    for ci, h in enumerate(vendor.columns, 1):
        ws_raw.cell(row=1, column=ci, value=h)
    _hdr_style(ws_raw, 1, len(vendor.columns), fill_hex="2c3e50")
    for ri, (_, row) in enumerate(vendor.iterrows(), start=2):
        for ci, val in enumerate(row, 1):
            ws_raw.cell(row=ri, column=ci, value=val)
    _alt_rows(ws_raw, 2, len(vendor)+1, len(vendor.columns))
    _auto_width(ws_raw)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_executive_summary_xlsx(vendor, kpis, report_month):
    wb = Workbook(); wb.remove(wb.active)
    build_executive_summary_sheet(wb, kpis, report_month)
    return _excel_bytes(wb)


def build_vendor_status_xlsx(vendor):
    wb = Workbook(); wb.remove(wb.active)
    build_vendor_status_sheet(wb, vendor)
    return _excel_bytes(wb)


def build_action_center_xlsx(action_df):
    wb = Workbook(); wb.remove(wb.active)
    build_action_center_sheet(wb, action_df)
    return _excel_bytes(wb)


# -----------------------------------------------------------------------------
# SIDEBAR  -- upload + month selector
# -----------------------------------------------------------------------------

def render_sidebar(conn):
    with st.sidebar:
        st.markdown(f"### {APP_TITLE}")
        st.markdown("---")
        st.markdown("**Upload CLM Export**")
        uploaded = st.file_uploader(
            "Excel (.xlsx) or CSV (.csv)",
            type=["xlsx","csv"],
            label_visibility="collapsed"
        )
        st.markdown("---")

        # Month selector -- shows all months stored in DB
        stored_months = get_all_stored_months(conn)
        selected_month = None
        if stored_months:
            st.markdown("**Select Report Month**")
            selected_month = st.selectbox(
                "Month",
                options=stored_months[::-1],   # latest first
                label_visibility="collapsed",
                key="selected_month"
            )
            st.markdown("---")
            st.markdown("**All Available Months**")
            for m in stored_months[::-1]:
                st.caption(f"- {m}")

        st.markdown("---")
        st.caption(
            "Data stored in local SQLite\n\n"
            "Original files never modified\n\n"
            "Each month is tracked independently"
        )
    return uploaded, selected_month


# -----------------------------------------------------------------------------
# TAB: DASHBOARD
# -----------------------------------------------------------------------------

def tab_dashboard(vendor, kpis, report_month, conn):
    st.markdown(
        f'<div class="section-header">Executive Dashboard - {report_month}</div>',
        unsafe_allow_html=True
    )

    c1,c2,c3,c4,c5,c6,c7 = st.columns(7)
    metrics = [
        (c1, "Total Vendors",     kpis["total"],    "blue"),
        (c2, "Total Employees",   kpis["emp"],      "blue"),
        (c3, "Compliant Vendors", kpis["comp"],     "green"),
        (c4, "Non-Compliant",     kpis["non"],      "red"),
        (c5, "Employees at Risk", kpis["risk"],     "amber"),
        (c6, "Compliance %",      f"{kpis['pct']}%",
             "green" if kpis["pct"] >= 80 else "amber"),
        (c7, "Critical Vendors",  kpis["critical"], "red"),
    ]
    for col, label, val, colour in metrics:
        col.markdown(
            f'<div class="kpi-card {colour}">'
            f'<div class="kpi-value">{val}</div>'
            f'<div class="kpi-label">{label}</div>'
            f'</div>',
            unsafe_allow_html=True
        )

    st.markdown("<br>", unsafe_allow_html=True)

    r1c1, r1c2 = st.columns(2)
    with r1c1:
        st.plotly_chart(chart_donut(kpis["comp"], kpis["non"]),
                        use_container_width=True, key="dash_donut")
    with r1c2:
        if not vendor[vendor["Health Score"] < 100].empty:
            st.plotly_chart(chart_employees_at_risk(vendor),
                            use_container_width=True, key="dash_emp_risk_top")
        else:
            st.success("All vendors compliant - no employees at risk!")

    r2c1, r2c2 = st.columns(2)
    with r2c1:
        pb_fig = chart_priority_breakdown(vendor)
        if pb_fig:
            st.plotly_chart(pb_fig, use_container_width=True,
                            key="dash_priority")
    with r2c2:
        rc_fig = chart_root_cause(vendor)
        if rc_fig:
            st.plotly_chart(rc_fig, use_container_width=True,
                            key="dash_root_cause")

    with st.expander("Region-wise Summary Table"):
        reg_sum = (
            vendor.groupby("Region").agg(
                Vendors=("Vendor Name","count"),
                Compliant=("Health Score", lambda x: (x==100).sum()),
                Non_Compliant=("Health Score", lambda x: (x<100).sum()),
                Employees=("No of Employees","sum"),
                At_Risk=("No of Employees",
                         lambda x: x[vendor.loc[x.index,"Health Score"]<100].sum()),
            ).reset_index()
        )
        reg_sum["Compliance %"] = (
            reg_sum["Compliant"] / reg_sum["Vendors"] * 100
        ).round(1).astype(str) + "%"
        st.dataframe(reg_sum, use_container_width=True, hide_index=True)


# -----------------------------------------------------------------------------
# TAB: VENDOR STATUS
# -----------------------------------------------------------------------------

def tab_vendor_status(vendor):
    st.markdown(
        '<div class="section-header">Vendor Status</div>',
        unsafe_allow_html=True
    )

    fc1, fc2, fc3, fc4 = st.columns([2,2,2,3])
    regions    = ["All"] + sorted(vendor["Region"].dropna().unique().tolist())
    statuses   = ["All","Healthy","Minor Issues","Needs Attention","Critical"]
    priorities = ["All","Critical","High","Medium","-"]

    region_sel   = fc1.selectbox("Region",   regions,    key="vs_region")
    status_sel   = fc2.selectbox("Status",   statuses,   key="vs_status")
    priority_sel = fc3.selectbox("Priority", priorities, key="vs_priority")
    search       = fc4.text_input("Search Vendor", key="vs_search")

    filtered = vendor.copy()
    if region_sel   != "All": filtered = filtered[filtered["Region"]==region_sel]
    if status_sel   != "All": filtered = filtered[filtered["Status"]==status_sel]
    if priority_sel != "All": filtered = filtered[filtered["Priority"]==priority_sel]
    if search:
        filtered = filtered[
            filtered["Vendor Name"].str.contains(search, case=False, na=False)
        ]

    st.caption(f"Showing **{len(filtered)}** of **{len(vendor)}** vendors")

    display_cols = [c for c in [
        "Vendor Name","V Code","Region","Location","City",
        "No of Employees","Active gatepass","Difference",
        "Health Score","Status","Priority","Issues",
        "Wage Compliance","PF Compliance","ESIC Compliance",
        "Reason for Non-Compliance"
    ] if c in filtered.columns]

    st.dataframe(
        filtered[display_cols].reset_index(drop=True),
        use_container_width=True, height=480,
    )

    buf = build_vendor_status_xlsx(filtered)
    st.download_button(
        "Download Filtered Vendor Status (Excel)",
        data=buf,
        file_name=f"vendor_status_{datetime.date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# -----------------------------------------------------------------------------
# TAB: ACTION CENTER
# -----------------------------------------------------------------------------

def tab_action_center(vendor, report_month, conn):
    st.markdown(
        '<div class="section-header">Action Center - Non-Compliant Vendors</div>',
        unsafe_allow_html=True
    )

    nc = vendor[vendor["Health Score"] < 100].copy()
    if nc.empty:
        st.success("All vendors are fully compliant. No actions required.")
        return

    tracking = get_tracking(conn, report_month)
    if not tracking.empty:
        nc = nc.merge(
            tracking.rename(columns={"vendor_name":"Vendor Name"}),
            on="Vendor Name", how="left"
        )

    for col in ["current_status","owner","remarks","last_updated"]:
        if col not in nc.columns:
            nc[col] = None

    nc["current_status"] = nc["current_status"].fillna("New")
    nc["owner"]          = nc["owner"].fillna("HR")
    nc["remarks"]        = nc["remarks"].fillna("")
    nc["last_updated"]   = nc["last_updated"].fillna("")
    nc["Status Aging"]   = nc["last_updated"].apply(compute_aging)

    ac1, ac2 = st.columns([2,3])
    priority_f = ac1.selectbox("Filter by Priority",
                               ["All","Critical","High","Medium"], key="ac_pri")
    status_f   = ac2.selectbox("Filter by Current Status",
                               ["All"] + STATUS_OPTIONS, key="ac_stat")

    view = nc.copy()
    if priority_f != "All": view = view[view["Priority"]==priority_f]
    if status_f   != "All": view = view[view["current_status"]==status_f]

    st.caption(f"**{len(view)}** vendor(s) requiring attention")
    st.markdown("---")
    st.markdown("**Update vendor status below - click Save after each vendor**")

    saved_any = False
    for _, row in view.sort_values(["Priority","Vendor Name"]).iterrows():
        vendor_name = row["Vendor Name"]
        prio_tag    = {"Critical":"[CRITICAL]","High":"[HIGH]",
                       "Medium":"[MEDIUM]","-":"[--]"}.get(row["Priority"],"[--]")

        with st.expander(
            f"{prio_tag} {vendor_name} | {row['Region']} | {row.get('City','-')} | "
            f"Score: {row['Health Score']} | {row['Priority']}"
        ):
            col_a, col_b = st.columns([3,2])
            with col_a:
                st.write(f"**Issues:** {row['Issues']}")
                st.write(f"**Recommendation:** {row['Recommended Action']}")
                rc_val = row.get("Reason for Non-Compliance","-")
                st.write(f"**Root Cause:** {rc_val}")
                lu_val = row['last_updated'] or 'Never'
                st.write(f"**Last Updated:** {lu_val} | **Aging:** {row['Status Aging']}")

            with col_b:
                kp = f"{report_month}_{vendor_name}".replace(" ","_")

                curr_status = st.selectbox(
                    "Current Status", STATUS_OPTIONS,
                    index=STATUS_OPTIONS.index(row["current_status"])
                          if row["current_status"] in STATUS_OPTIONS else 0,
                    key=f"status_{kp}"
                )
                curr_owner = st.selectbox(
                    "Owner", OWNER_OPTIONS,
                    index=OWNER_OPTIONS.index(row["owner"])
                          if row["owner"] in OWNER_OPTIONS else 0,
                    key=f"owner_{kp}"
                )
                curr_remarks = st.text_area(
                    "Remarks", value=row["remarks"],
                    key=f"remarks_{kp}", height=80
                )

                if st.button("Save", key=f"save_{kp}"):
                    upsert_tracking(conn, report_month, vendor_name,
                                    curr_status, curr_owner, curr_remarks)
                    st.success(f"Saved - {vendor_name}")
                    saved_any = True

    if saved_any:
        st.rerun()

    st.markdown("---")
    ac_cols = ["Vendor Name","Region","City","Issues","Priority",
               "current_status","owner","remarks","last_updated",
               "Status Aging","Recommended Action"]
    ac_dl = view[[c for c in ac_cols if c in view.columns]].rename(columns={
        "current_status":"Current Status","owner":"Owner",
        "remarks":"Remarks","last_updated":"Last Updated",
    })
    buf = build_action_center_xlsx(ac_dl)
    st.download_button(
        "Download Action Center (Excel)",
        data=buf,
        file_name=f"action_center_{report_month.replace(' ','_')}_{datetime.date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# -----------------------------------------------------------------------------
# TAB: ANALYTICS
# -----------------------------------------------------------------------------

def tab_analytics(vendor, report_month):
    st.markdown(
        '<div class="section-header">Analytics and Deep Dive</div>',
        unsafe_allow_html=True
    )

    fig2 = px.scatter(vendor, x="No of Employees", y="Health Score",
                      color="Status", size="No of Employees",
                      hover_name="Vendor Name",
                      color_discrete_map={
                          "Healthy":"#27ae60","Minor Issues":"#f39c12",
                          "Needs Attention":"#e67e22","Critical":"#e74c3c"
                      },
                      title="Health Score vs Employee Count")
    fig2.update_layout(height=380, margin=dict(t=40,b=10,l=10,r=10),
                       paper_bgcolor="white", plot_bgcolor="white",
                       font=dict(color="#1a1a1a", size=12),
                       title_font=dict(color="#1a1a1a", size=14),
                       legend=dict(font=dict(color="#1a1a1a")),
                       xaxis=dict(tickfont=dict(color="#1a1a1a"), title_font=dict(color="#1a1a1a")),
                       yaxis=dict(tickfont=dict(color="#1a1a1a"), title_font=dict(color="#1a1a1a")))
    st.plotly_chart(fig2, use_container_width=True, key="analytics_scatter")

    st.markdown("**Component-wise Compliance Analysis**")
    st.plotly_chart(chart_component_compliance(vendor),
                    use_container_width=True, key="analytics_component")

    rc_fig = chart_root_cause(vendor)
    if rc_fig:
        st.plotly_chart(rc_fig, use_container_width=True,
                        key="analytics_root_cause")

    with st.expander("Full Vendor Analytics Table"):
        st.dataframe(
            vendor[[c for c in [
                "Vendor Name","Region","Location","City",
                "No of Employees","Health Score","Status","Priority",
                "Wage Compliance","PF Compliance","ESIC Compliance","Difference"
            ] if c in vendor.columns]].sort_values("Health Score"),
            use_container_width=True, hide_index=True
        )


# -----------------------------------------------------------------------------
# TAB: REPORTS
# -----------------------------------------------------------------------------

def tab_reports(vendor, kpis, report_month, conn):
    st.markdown(
        '<div class="section-header">Reports and Downloads</div>',
        unsafe_allow_html=True
    )
    st.info(
        "All reports are generated fresh from current data. "
        "The original uploaded file is never modified."
    )

    nc = vendor[vendor["Health Score"] < 100].copy()
    tracking = get_tracking(conn, report_month)
    if not tracking.empty and not nc.empty:
        nc = nc.merge(
            tracking.rename(columns={"vendor_name":"Vendor Name"}),
            on="Vendor Name", how="left"
        )
    for col in ["current_status","owner","remarks","last_updated"]:
        if col not in nc.columns:
            nc[col] = "-"
    nc["current_status"] = nc["current_status"].fillna("New")
    nc["Status Aging"]   = nc.get("last_updated","").apply(compute_aging)
    ac_dl = nc.rename(columns={
        "current_status":"Current Status","owner":"Owner",
        "remarks":"Remarks","last_updated":"Last Updated",
    })

    st.markdown("---")
    r1, r2 = st.columns(2)
    with r1:
        st.markdown("#### Executive Summary")
        st.caption("KPI snapshot for management presentation.")
        st.download_button(
            "Download Executive Summary",
            data=build_executive_summary_xlsx(vendor, kpis, report_month),
            file_name=f"executive_summary_{report_month.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with r2:
        st.markdown("#### Vendor Status Report")
        st.caption("Full vendor list with health scores and compliance status.")
        st.download_button(
            "Download Vendor Status",
            data=build_vendor_status_xlsx(vendor),
            file_name=f"vendor_status_{report_month.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    r3, r4 = st.columns(2)
    with r3:
        st.markdown("#### Action Center Report")
        st.caption("Non-compliant vendors with current status, owner, and remarks.")
        st.download_button(
            "Download Action Center",
            data=build_action_center_xlsx(ac_dl),
            file_name=f"action_center_{report_month.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with r4:
        st.markdown("#### Complete Report (All Sheets)")
        st.caption("Executive Summary + Vendor Status + Action Center + Raw Data.")
        st.download_button(
            "Download Complete Report",
            data=build_complete_report(vendor, kpis, ac_dl, report_month),
            file_name=f"complete_report_{report_month.replace(' ','_')}_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    st.markdown("#### Audit Log")
    audit = pd.read_sql_query(
        "SELECT ts, report_month, vendor_name, field, old_value, new_value "
        "FROM AuditLog ORDER BY ts DESC LIMIT 200",
        conn
    )
    if audit.empty:
        st.caption("No changes recorded yet.")
    else:
        st.dataframe(audit, use_container_width=True, height=280)
        st.download_button(
            "Download Audit Log (CSV)",
            data=audit.to_csv(index=False).encode(),
            file_name="audit_log.csv",
            mime="text/csv"
        )


# -----------------------------------------------------------------------------
# TAB: SETTINGS
# -----------------------------------------------------------------------------

def tab_settings(conn):
    st.markdown(
        '<div class="section-header">Settings and Administration</div>',
        unsafe_allow_html=True
    )

    st.markdown("#### Upload History")
    hist = get_upload_history(conn)
    if hist.empty:
        st.caption("No uploads yet.")
    else:
        st.dataframe(hist, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("#### Business Rules Reference")
    st.markdown("""
| Component | Score | Non-Complied Priority |
|---|---|---|
| Wage Compliance | +25 | Critical |
| PF Compliance | +25 | Critical |
| ESIC Compliance | +25 | Critical |
| Gatepass Difference = 0 | +25 | High |

**Health Score Bands:**
- 100 - Healthy
- 75 to 99 - Minor Issues
- 50 to 74 - Needs Attention
- 0 to 49 - Critical
    """)

    st.markdown("---")
    st.markdown("#### Required Columns")
    st.code("\n".join(REQUIRED_COLUMNS))

    st.markdown("---")
    st.markdown("#### Clear Tracking Data")
    st.warning("This will delete ALL status updates. This cannot be undone.")
    if st.button("Clear ALL Tracking Data", type="secondary"):
        conn.execute("DELETE FROM TrackingData")
        conn.execute("DELETE FROM AuditLog")
        conn.commit()
        st.success("Tracking data cleared.")
        st.rerun()

    st.markdown("---")
    st.markdown("#### Clear Monthly Stored Data")
    st.warning("This will delete all stored monthly vendor snapshots. You will need to re-upload your Excel files.")
    if st.button("Clear Monthly Vendor Snapshots", type="secondary"):
        conn.execute("DELETE FROM MonthlyVendorData")
        conn.execute("DELETE FROM UploadHistory")
        conn.commit()
        st.success("Monthly data cleared.")
        st.rerun()


# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------

def main():
    conn = get_db()
    uploaded, selected_month = render_sidebar(conn)

    # ── Landing page ──
    if uploaded is None and not get_all_stored_months(conn):
        st.markdown(f"## {APP_TITLE}")
        st.markdown("---")
        col_l, col_r = st.columns([2,1])
        with col_l:
            st.markdown("""
### Welcome to the HR Contractor Cell Control Tower

This tool helps your HR team:

- **Monitor** vendor compliance status at a glance
- **Track** every issue with owner, status, and aging
- **Select any month** independently - no data mixing
- **Report** to management with one-click Excel exports

#### How to use:
1. Export data from your CLM software (Excel or CSV)
2. Upload it using the sidebar
3. All monthly sheets are read and stored separately
4. Use the month selector in the sidebar to switch months
5. Update vendor statuses in the Action Center per month
6. Download reports for any selected month

#### What data is needed:
Your CLM export must include columns for Region, Location,
Vendor Name, Employee headcount, Gatepass data, and
Wage / PF / ESIC compliance status.
            """)
        with col_r:
            st.markdown("#### Quick Stats")
            hist = get_upload_history(conn)
            if not hist.empty:
                st.metric("Total Months Stored", len(hist))
                st.metric("Latest Month", hist.iloc[0]["report_month"])
            else:
                st.info("No data uploaded yet. Upload your CLM Excel file to begin.")
        st.stop()

    # ── Process new upload ──
    if uploaded is not None:
        with st.spinner("Reading file and processing all monthly sheets..."):
            month_data, sorted_months = parse_upload(uploaded)

        # Validate and store each month
        for month, raw_df in month_data.items():
            missing = validate_columns(raw_df)
            if missing:
                st.error(
                    f"Sheet '{month}' - Validation Failed. Missing columns:\n\n"
                    + "\n".join(f"- `{m}`" for m in missing)
                )
                with st.expander(f"Columns found in sheet '{month}'"):
                    st.write(raw_df.columns.tolist())
                continue

            vendor_df = build_vendor_df(raw_df)
            save_monthly_vendor_data(conn, month, vendor_df)
            log_upload(conn, month, len(raw_df), len(vendor_df))

        n = len(month_data)
        st.success(
            f"Upload complete. {n} monthly sheet(s) processed and stored: "
            f"{', '.join(sorted_months)}. "
            "Use the month selector in the sidebar to view each month."
        )

    # ── Load selected month data ──
    stored_months = get_all_stored_months(conn)
    if not stored_months:
        st.warning("No data available. Please upload a CLM export file.")
        st.stop()

    # Default to latest if nothing selected yet
    if selected_month is None or selected_month not in stored_months:
        selected_month = stored_months[-1]

    vendor = load_monthly_vendor_data(conn, selected_month)
    if vendor is None:
        st.error(f"Could not load data for {selected_month}. Please re-upload.")
        st.stop()

    # Ensure numeric columns are correct dtype after JSON round-trip
    for col in ["No of Employees","Active gatepass","Difference","Health Score"]:
        if col in vendor.columns:
            vendor[col] = pd.to_numeric(vendor[col], errors="coerce").fillna(0)

    kpis = compute_kpis(vendor)

    # Month banner
    st.markdown(
        f'<div class="month-banner">'
        f'Viewing: <strong>{selected_month}</strong> | '
        f'Vendors: {kpis["total"]} | '
        f'Compliance: {kpis["pct"]}% | '
        f'Non-Compliant: {kpis["non"]} | '
        f'Critical: {kpis["critical"]}'
        f'</div>',
        unsafe_allow_html=True
    )

    tabs = st.tabs([
        "Dashboard",
        "Vendor Status",
        "Action Center",
        "Analytics",
        "Reports",
        "Settings",
    ])

    with tabs[0]: tab_dashboard(vendor, kpis, selected_month, conn)
    with tabs[1]: tab_vendor_status(vendor)
    with tabs[2]: tab_action_center(vendor, selected_month, conn)
    with tabs[3]: tab_analytics(vendor, selected_month)
    with tabs[4]: tab_reports(vendor, kpis, selected_month, conn)
    with tabs[5]: tab_settings(conn)


if __name__ == "__main__":
    main()
